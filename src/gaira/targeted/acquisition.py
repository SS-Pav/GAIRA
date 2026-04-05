"""High-precision targeted acquisition for Raman/SERS files."""

from __future__ import annotations

from hashlib import sha256
from io import BytesIO, StringIO
from pathlib import Path
from urllib.parse import parse_qs, quote_plus, unquote, urljoin, urlparse
import csv
import re
import zipfile

from bs4 import BeautifulSoup
import requests

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


SEARCH_ENDPOINT = "https://duckduckgo.com/html/?q={query}"
REQUEST_TIMEOUT = 30
MAX_SEARCH_RESULTS = 50
MAX_FILE_BYTES = 50 * 1024 * 1024
SEARCH_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; GAIRA_TARGETED_ACQUISITION/1.0; "
        "+https://github.com/openai)"
    )
}
HIGH_VALUE_DOMAINS = (
    "zenodo.org",
    "figshare.com",
    "github.com",
    ".edu",
    "mdpi.com",
    "nature.com",
    "sciencedirect.com",
    "rsc.org",
    "acs.org",
    "data.mendeley.com",
)
HIGH_VALUE_URL_TERMS = ("supplementary", "dataset", "download", "csv", "xlsx")
FILE_EXTENSIONS = (".csv", ".xlsx", ".txt", ".zip", ".tsv")
TEXT_PRIORITY_TERMS = ("raman", "sers")
SUPPLEMENTARY_URL_TERMS = ("supp", "supplement", "supporting")

def generate_search_queries() -> list[str]:
    return [
        "Raman spectra csv serum",
        "SERS spectra dataset download",
        "Raman metabolite spectra csv",
        "Raman spectral library biomolecules",
        "SERS exosome spectra dataset",
        "Raman extracellular vesicles spectra",
        "Raman glycan spectra csv",
        "Raman lipid spectra dataset",
        "Raman protein spectra csv",
        "Raman supplementary data spectra filetype:xlsx",
        "SERS dataset filetype:csv",
        "Raman spectra filetype:txt",
        "Raman spectra filetype:xlsx",
        "SERS supplementary material spectra",
    ]


def _request_text(url: str) -> str:
    response = requests.get(url, timeout=REQUEST_TIMEOUT, headers=SEARCH_HEADERS)
    response.raise_for_status()
    return response.text


def request_json(url: str, params: dict[str, object] | None = None) -> dict[str, object]:
    response = requests.get(url, params=params, timeout=REQUEST_TIMEOUT, headers=SEARCH_HEADERS)
    response.raise_for_status()
    return response.json()


def request_xml_text(url: str, params: dict[str, object] | None = None) -> str:
    response = requests.get(url, params=params, timeout=REQUEST_TIMEOUT, headers=SEARCH_HEADERS)
    response.raise_for_status()
    return response.text


def request_post_json(
    url: str,
    json_payload: dict[str, object] | None = None,
    params: dict[str, object] | None = None,
) -> object:
    response = requests.post(
        url,
        json=json_payload,
        params=params,
        timeout=REQUEST_TIMEOUT,
        headers=SEARCH_HEADERS,
    )
    response.raise_for_status()
    return response.json()


def _request_stream(url: str) -> requests.Response:
    response = requests.get(
        url,
        timeout=REQUEST_TIMEOUT,
        headers=SEARCH_HEADERS,
        stream=True,
        allow_redirects=True,
    )
    response.raise_for_status()
    return response


def _slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")[:80] or "query"


def _url_hash(value: str) -> str:
    return sha256(value.encode("utf-8")).hexdigest()[:16]


def _save_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _normalize_search_result_url(href: str) -> str:
    if href.startswith("//"):
        href = f"https:{href}"
    if href.startswith("/"):
        href = urljoin("https://duckduckgo.com", href)

    parsed = urlparse(href)
    if "duckduckgo.com" in parsed.netloc and parsed.path.startswith("/l/"):
        uddg_values = parse_qs(parsed.query).get("uddg", [])
        if uddg_values:
            return unquote(uddg_values[0])
    return href


def _is_duckduckgo_anomaly_page(html: str) -> bool:
    lowered = html.lower()
    return "anomaly-modal" in lowered or "bots use duckduckgo too" in lowered


def _extract_search_results(html: str, query: str) -> list[dict[str, str]]:
    if _is_duckduckgo_anomaly_page(html):
        return []
    soup = BeautifulSoup(html, "html.parser")
    results: list[dict[str, str]] = []
    seen_urls: set[str] = set()

    for anchor in soup.select("a.result__a"):
        href = _normalize_search_result_url(anchor.get("href", "").strip())
        title = anchor.get_text(" ", strip=True)
        if not href or href in seen_urls:
            continue
        seen_urls.add(href)
        results.append({"query": query, "title": title, "url": href})
        if len(results) >= MAX_SEARCH_RESULTS:
            break

    return results


def search_duckduckgo_html(query: str, raw_html_dir: Path) -> list[dict[str, str]]:
    url = SEARCH_ENDPOINT.format(query=quote_plus(query))
    html = _request_text(url)
    raw_path = raw_html_dir / f"search_{_slugify(query)}.html"
    _save_text(raw_path, html)
    return _extract_search_results(html, query)


def search_page_has_anomaly(query: str, raw_html_dir: Path) -> bool:
    raw_path = raw_html_dir / f"search_{_slugify(query)}.html"
    if not raw_path.exists():
        return False
    return _is_duckduckgo_anomaly_page(raw_path.read_text(encoding="utf-8"))


def _is_high_value_url(url: str) -> bool:
    lowered = url.lower()
    parsed = urlparse(lowered)
    host = parsed.netloc
    has_domain_match = any(domain in host for domain in HIGH_VALUE_DOMAINS)
    has_term_match = any(term in lowered for term in HIGH_VALUE_URL_TERMS)
    return has_domain_match or has_term_match


def filter_target_urls(results: list[dict[str, str]]) -> list[dict[str, str]]:
    filtered: list[dict[str, str]] = []
    seen_urls: set[str] = set()
    for result in results:
        url = result.get("url", "")
        if url in seen_urls:
            continue
        if _is_high_value_url(url):
            filtered.append(result)
            seen_urls.add(url)
    return filtered


def _normalize_page_links(base_url: str, html: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    links: list[str] = []
    seen_urls: set[str] = set()
    for anchor in soup.find_all("a", href=True):
        absolute_url = urljoin(base_url, anchor["href"]).strip()
        if absolute_url and absolute_url not in seen_urls:
            links.append(absolute_url)
            seen_urls.add(absolute_url)
    return links


def _is_candidate_file_link(url: str) -> bool:
    lowered = url.lower()
    return lowered.endswith(FILE_EXTENSIONS) or "supplementary" in lowered


def _page_priority_boost(text: str) -> bool:
    lowered = text.lower()
    return all(term in lowered for term in TEXT_PRIORITY_TERMS)


def discover_page_file_links(result: dict[str, str], raw_html_dir: Path) -> tuple[list[str], bool]:
    page_url = result.get("url", "")
    html = _request_text(page_url)
    page_path = raw_html_dir / f"page_{_url_hash(page_url)}.html"
    _save_text(page_path, html)
    links = _normalize_page_links(page_url, html)
    file_links = [link for link in links if _is_candidate_file_link(link)]
    priority_boost = _page_priority_boost(html)
    return file_links, priority_boost


def _infer_file_name(url: str) -> str:
    parsed = urlparse(url)
    name = Path(parsed.path).name
    if name:
        return name
    return f"download_{_url_hash(url)}.bin"


def _infer_file_type(file_name: str) -> str:
    suffix = Path(file_name).suffix.lower().lstrip(".")
    return suffix or "unknown"


def prioritize_supplementary_url(url: str) -> bool:
    lowered = url.lower()
    return any(term in lowered for term in SUPPLEMENTARY_URL_TERMS)


def initialize_content_hashes(files_dir: Path) -> set[str]:
    hashes: set[str] = set()
    if not files_dir.exists():
        return hashes
    for path in files_dir.iterdir():
        if not path.is_file():
            continue
        try:
            hashes.add(sha256(path.read_bytes()).hexdigest())
        except OSError:
            continue
    return hashes


def download_candidate_file(
    file_url: str,
    source_url: str,
    files_dir: Path,
    priority_boost: bool,
    source_type: str,
    extraction_method: str,
    content_hashes: set[str] | None = None,
) -> dict[str, object] | None:
    response = _request_stream(file_url)
    content_length = response.headers.get("Content-Length")
    if content_length:
        try:
            if int(content_length) > MAX_FILE_BYTES:
                response.close()
                return None
        except ValueError:
            pass

    file_name = _infer_file_name(file_url)
    file_hash = _url_hash(file_url)
    target_name = f"{file_hash}_{file_name}"
    target_path = files_dir / target_name
    if target_path.exists():
        existing_hash = ""
        try:
            existing_hash = sha256(target_path.read_bytes()).hexdigest()
            if content_hashes is not None:
                content_hashes.add(existing_hash)
        except OSError:
            existing_hash = ""
        return {
            "file_name": target_name,
            "source_url": source_url,
            "file_type": _infer_file_type(file_name),
            "size_kb": max(1, int(target_path.stat().st_size / 1024)),
            "local_path": target_path,
            "priority_boost": priority_boost,
            "source_type": source_type,
            "extraction_method": extraction_method,
            "content_hash": existing_hash,
        }

    bytes_written = 0
    digest = sha256()
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with target_path.open("wb") as handle:
        for chunk in response.iter_content(chunk_size=8192):
            if not chunk:
                continue
            bytes_written += len(chunk)
            if bytes_written > MAX_FILE_BYTES:
                handle.close()
                target_path.unlink(missing_ok=True)
                response.close()
                return None
            handle.write(chunk)
            digest.update(chunk)

    response.close()
    content_hash = digest.hexdigest()
    if content_hashes is not None and content_hash in content_hashes:
        target_path.unlink(missing_ok=True)
        return None
    if content_hashes is not None:
        content_hashes.add(content_hash)

    return {
        "file_name": target_name,
        "source_url": source_url,
        "file_type": _infer_file_type(file_name),
        "size_kb": max(1, int(bytes_written / 1024)),
        "local_path": target_path,
        "priority_boost": priority_boost,
        "source_type": source_type,
        "extraction_method": extraction_method,
        "content_hash": content_hash,
    }


def _parse_numeric_rows(rows: list[list[object]]) -> tuple[bool, bool]:
    numeric_pairs: list[tuple[float, float]] = []
    for row in rows:
        if len(row) < 2:
            continue
        first = _to_float(row[0])
        second = _to_float(row[1])
        if first is None or second is None:
            continue
        numeric_pairs.append((first, second))
        if len(numeric_pairs) >= 25:
            break

    if len(numeric_pairs) < 5:
        return False, False

    increasing = all(
        numeric_pairs[index][0] < numeric_pairs[index + 1][0]
        for index in range(len(numeric_pairs) - 1)
    )
    return True, increasing


def _contains_table_terms(rows: list[list[object]]) -> bool:
    for row in rows[:10]:
        combined = " ".join(str(cell).lower() for cell in row if cell is not None)
        if "peak" in combined or "cm-1" in combined or "cm^-1" in combined:
            return True
    return False


def _classify_rows(rows: list[list[object]]) -> str:
    if _contains_table_terms(rows):
        return "TABLE"

    has_numeric_rows, increasing = _parse_numeric_rows(rows)
    if has_numeric_rows and increasing:
        return "RAW_SPECTRA"

    return "UNKNOWN"


def _to_float(value: object) -> float | None:
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def _read_delimited_rows(file_obj: StringIO, delimiter: str) -> list[list[object]]:
    reader = csv.reader(file_obj, delimiter=delimiter)
    rows: list[list[object]] = []
    for index, row in enumerate(reader):
        rows.append(row)
        if index >= 50:
            break
    return rows


def _classify_text_buffer(buffer: bytes, file_type: str) -> str:
    decoded = buffer.decode("utf-8", errors="ignore")
    delimiter = "\t" if file_type == "tsv" else ","
    rows = _read_delimited_rows(StringIO(decoded), delimiter)
    return _classify_rows(rows)


def _classify_xlsx_buffer(buffer: bytes) -> str:
    if load_workbook is None:
        return "UNKNOWN"
    workbook = load_workbook(BytesIO(buffer), read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[list[object]] = []
    for index, row in enumerate(worksheet.iter_rows(values_only=True)):
        rows.append(list(row))
        if index >= 50:
            break
    workbook.close()
    return _classify_rows(rows)


def _classify_zip_file(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as archive:
            for member in archive.namelist():
                lowered = member.lower()
                if lowered.endswith((".csv", ".txt", ".tsv", ".xlsx")):
                    with archive.open(member) as handle:
                        buffer = handle.read()
                    if lowered.endswith(".xlsx"):
                        result = _classify_xlsx_buffer(buffer)
                    elif lowered.endswith(".tsv"):
                        result = _classify_text_buffer(buffer, "tsv")
                    else:
                        result = _classify_text_buffer(buffer, "csv")
                    if result != "UNKNOWN":
                        return result
    except zipfile.BadZipFile:
        return "UNKNOWN"
    return "UNKNOWN"


def classify_downloaded_file(path: Path, file_type: str) -> str:
    try:
        if file_type in {"csv", "txt"}:
            return _classify_text_buffer(path.read_bytes(), "csv")
        if file_type == "tsv":
            return _classify_text_buffer(path.read_bytes(), "tsv")
        if file_type == "xlsx":
            return _classify_xlsx_buffer(path.read_bytes())
        if file_type == "zip":
            return _classify_zip_file(path)
    except Exception:
        return "UNKNOWN"
    return "UNKNOWN"


def write_registry(records: list[dict[str, object]], output_path: Path) -> None:
    fieldnames = [
        "file_name",
        "source_url",
        "file_type",
        "spectra_type",
        "size_kb",
        "source_type",
        "extraction_method",
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow({field: record.get(field, "") for field in fieldnames})
